import openpyxl
import pandas as pd

class ExcelFile:

    def __init__(self, file_name):
        self.filename = file_name
        self.workbook = openpyxl.load_workbook(self.filename)
        self.worksheet = self.workbook.active

    def close_file(self):
        self.workbook.close()

    def countRows(self):
        # Count rows
        return self.worksheet.max_row
    
    def countColumns(self):
        # Count columns
        return self.worksheet.max_column

    def writeOnXL(self, data):

        self.update_excel(self.countRows() - 1, data)

    # ---> FOR SO_FORM <--- #
    def read_into_dataframe_SO(self):
        # Read the Excel file into a pandas DataFrame
        return pd.read_excel(self.filename,  dtype={'PHONE': str, 'ZIPCODE': str})

    # ---> FOR AI_FORM <--- #
    def read_into_dataframe(self):
        # Read the Excel file into a pandas DataFrame
        return pd.read_excel(self.filename, 'MASTER')
    
    def read_into_dataframe_ai(self):
        # Read the Excel file into a pandas DataFrame
        return pd.read_excel(self.filename, 'Sheet1')
    
    def update_excel(self, row, new_values):
        row += 2
        # Update the Excel file with the new values
        try:
            for column_title, value in new_values.items():
                column_index = self.get_column_index(column_title)
                self.worksheet.cell(row=row, column=column_index, value=value)
            self.workbook.save(self.filename)
            self.close_file()
        except AttributeError as e:
            print("AttributeError: " + str(new_values))

    def get_column_index(self, column_title):
        # Get the column index based on the column title
        for column in self.worksheet.iter_cols(min_row=1, max_row=1):
            for cell in column:
                if cell.value == column_title:
                    return cell.column
        raise ValueError(f"Column '{column_title}' not found in the Excel file.")
    
    def new_row(self, date):
        if self.get_todays_entries_count(date) == 1:
            row = self.countRows()
            for i in range(1, self.countColumns() + 1):
                self.worksheet.cell(row=row+ 1, column=i).value = " "
            self.workbook.save(self.filename)
            self.close_file()

    def get_column_names(self):
        # Get the column names
        return [cell.value for cell in self.worksheet[1]]
    
    def get_todays_entries_count(self, today):
        # Set count to 1 to avoid triple 0 entry
        count = 1
        for row in range(1, self.worksheet.max_row + 1):
            if self.worksheet.cell(row=row, column=1).value == today:
                count += 1
        return count
    
    def get_crates(self):

        df = pd.read_excel(self.filename, 'MASTER')

        # Drop NA values before doing string operations
        df.dropna(subset=['Assignment'], inplace=True)

        # Gather all crates that have "Crate #" in the 'Assignment' column
        df = df[df['Assignment'].str.contains("Crate")]

        # Extract the crate numbers from the 'Assignment' column
        crate_numbers = df['Assignment'].unique().tolist()

        if not crate_numbers:
            crate_numbers.sort()

        ff = pd.read_excel(self.filename, 'MASTER')

        # Drop NA values before doing string operations
        ff.dropna(subset=['Assignment'], inplace=True)

        # Obtain the crate numbers from the 'Assignment' column that contain "RSL", filtering the rows with the same first 6 digits after 'RSL'
        df_rsl = ff[ff['Assignment'].str.contains("RSL") & ~ff['Assignment'].str.contains("P-RSL")]
        df_rsl.loc[:, 'Assignment'] = df_rsl['Assignment'].str[:9]

        #print("rsl's: " + str(df_rsl['Assignment'].unique().tolist()))

        # add to list of crates
        crate_numbers.extend(df_rsl['Assignment'].unique().tolist())

        return crate_numbers
    
    def filter_by_crate_num(self, crate_num):
        df = pd.read_excel(self.filename, 'MASTER')
        df = df[df['Assignment'] == "Crate " + str(crate_num)]
        return df
    
    # get the row indexes for the current crate number that do not have 'Con' and 'Notes' filled out
    def get_row_indexes_needing_highlights(self, crate_num):
        df = self.filter_by_crate_num(crate_num)

        # Convert values in the "Qty" column to strings
        df["Qty"] = df["Qty"].astype(str)

        # Filter rows to get the ones where 'Con' contains "Highlights"
        mask = df["Qty"].str.contains("Highlights", na=False)
        df = df[mask]
        
        idxs = df.index.tolist()
        
        if not idxs:
            idxs.sort()

        return idxs
    
    # get the row indexes for the current crate number that does have 'Con' filled out and have 1 year in 'Notes'
    def get_row_indexes_needing_grading(self, crate_num):
        
        df = self.filter_by_crate_num(crate_num)

        # Convert values in the "Qty" column to strings
        df["Qty"] = df["Qty"].astype(str)

        # Filter rows to get the ones where 'Con' contains "Grading"
        mask = df["Qty"].str.contains("Grading", na=False)
        df = df[mask]

        idxs = df.index.tolist()
        
        if not idxs:
            idxs.sort()

        return idxs
    
    def get_row_indexes_needing_intro(self, crate_num):
        
        df = self.filter_by_crate_num(crate_num)

        # Convert values in the "Qty" column to strings
        df["Qty"] = df["Qty"].astype(str)

        # Filter rows to get the ones where 'Con' contains "Grading"
        mask = df["Qty"].str.contains("Intro", na=False)
        df = df[mask]

        idxs = df.index.tolist()
        
        if not idxs:
            idxs.sort()

        return idxs
    
    def get_row_indexes_needing_RSL(self, crate_num):
        df = self.filter_by_crate_num(crate_num)

        conditions_list = [
            "New", "Very Good", "Good", "Fairly Good", "Fair"
        ]

        # Filter out the rows that the 'Assignment' string is in the conditions list
        mask = ~df['Assignment'].isin(conditions_list)
        df = df[mask]

        idxs = df.index.tolist()
        
        if not idxs:
            idxs.sort()

        return idxs
    
    def get_row_indexes_needing_printing(self, date):
        df = pd.read_excel(self.filename, 'MASTER')

        df.dropna(subset=['Assignment'], inplace=True)

        
        # Filter out the rows that the string in 'Assignment' contains "RSL" in the beginning but does not have "P-"
        mask = df['Assignment'].str.contains(date) & ~df['Assignment'].str.contains("P-")
        df = df[mask]        

        idxs = df.index.tolist()

        return idxs
    
    
    def crates_list_for_hightlights(self):
        df = pd.read_excel(self.filename, 'MASTER')

        # Drop rows where 'Assignment' is not filled out
        df = df.dropna(subset=['Assignment'])

        # Drop rows where 'Cond' and 'Notes' are filled out
        mask = df[["Qty"]].notna().any(axis=1)
        df = df[~mask]

        # Filter out rows where 'Assignment' is not filled out
        df = df.dropna(subset=['Assignment'])

        # Filter results with 'Notes' column length <= 5
        df['Notes'] = df['Notes'].astype(str)
        df = df[df['Notes'].str.len() <= 5]

        # Extract the crate numbers from the 'Assignment' column
        crate_numbers = df['Assignment'].unique().tolist()
        
        if not crate_numbers:
            crate_numbers.sort()

        return crate_numbers
    
    def crates_list_for_ai(self):
        df = pd.read_excel(self.filename, 'MASTER')

        # Drop rows where 'Assignment' is not filled out
        df = df.dropna(subset=['Assignment'])

        # Drop rows where 'Cond' and 'Notes' are filled out
        mask = df[["Qty"]].notna().any(axis=1)
        df = df[~mask]

        # Filter out rows where 'Assignment' is not filled out
        df = df.dropna(subset=['Assignment'])

        # Filter results with 'Notes' column length >= 5
        df['Notes'] = df['Notes'].astype(str)
        df = df[df['Notes'].str.len() > 5]

        # Extract the crate numbers from the 'Assignment' column
        crate_numbers = df['Assignment'].unique().tolist()
        
        if not crate_numbers:
            crate_numbers.sort()

        return crate_numbers
    

    def is_crate_there(self, crate_num):
        df = pd.read_excel(self.filename, 'MASTER')
        df = df[df['Assignment'] == "Crate " + str(crate_num)]
        if df.empty:
            return -1
        else:
            return df.index[-1] + 2
    
    def insert_row(self, row_number, data=None):
        # Insert a new row at the specified row number
        self.worksheet.insert_rows(row_number)
        if data:
            for column_title, value in data.items():
                column_index = self.get_column_index(column_title)
                self.worksheet.cell(row=row_number, column=column_index, value=value)
        self.workbook.save(self.filename)
        self.close_file()
    
    def refresh_excel_file(self):
        self.workbook = openpyxl.load_workbook(self.filename)
        self.worksheet = self.workbook['MASTER']



def open_excel_file(file_name): 
    return ExcelFile(file_name)
