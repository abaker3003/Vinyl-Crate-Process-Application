import threading
from tkinter import ttk
import customtkinter as ctk
import datetime
import os
import glob
from openpyxl import load_workbook
import pandas as pd
import xlfile as xl
import desc_input_frame as dif
import hightlights as II
import third_section as III
import progress_bar as prog
import rsl_price_frame as rsl

ctk.set_default_color_theme("red.json")

class Main(ctk.CTk):

    def __init__(self):
        super().__init__()
        
        self.state('zoomed')
        self.resizable(False, False)

        '''screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        self.geometry(f"{screen_width}x{screen_height}")'''
        
        '''
            D:\\ For Ashley's Computer
            E:\\ For George's Computer
            F:\\ For All Business' Computers
        '''
        #os.chdir("D:\\LP Prep Files")
        self.destination = "D:\\RSL\\RSL Load.xlsx"
        os.chdir("D:\\LP TEST")
        self.start_app()

    def start_app(self):
        self.new_crate = False
        self.title("All That Music & Video - Crate Process")

        self._last_frame = None

        self.files = os.listdir()

        self.file_crate_dict = {}
        self.current_crate_num = 0
        self.vinyl_data = {}
        self.file_crate_dict = {}
        self.existing_crates = []
        
        self.display_files()
    
    @property
    def last_frame(self):
        return self._last_frame
    
    @last_frame.setter
    def last_frame(self, last_frame):
        self._last_frame = last_frame

    def display_files(self):
        self.clear()
        self.show_crates()

    def show_crates(self):
        
        self.crates = []

        for i, file in enumerate(self.files):
            xl_handle = xl.open_excel_file(file)
            self.file_crate_dict[file] = xl_handle.get_crates()
            self.crates.extend(self.file_crate_dict[file])
        
        self.choose_file_frame = ctk.CTkFrame(self)
        self.choose_file_frame.grid(row=0, column=0, sticky='nsew', padx=10, pady=10)

        self.choose_file_frame.rowconfigure(0, weight=1)
        self.choose_file_frame.columnconfigure(1, weight=1)

        self.choose_file_label = ctk.CTkLabel(self.choose_file_frame,
                                               text="Choose a crate", font=("Courier New Greek", 18))
        self.choose_file_label.grid(row=0, column=0, columnspan=2, pady=20, sticky='nsew', padx=10)
        
        btn = ctk.CTkButton(self.choose_file_frame, text="New Crate", command=self.create_new_crate, font=("Courier New Greek", 18))
        btn.grid(row=1, column=0, pady=20, sticky='nsew', padx=10)

        self.choose_file_frame_scrollable = ctk.CTkScrollableFrame(self.choose_file_frame, orientation="vertical", width=720, height=450)
        self.choose_file_frame_scrollable.grid(row=2, column=0, padx=10, pady=10)
        
        if len(self.crates) != 0:
            def number(val):
                return int(val.split()[1])
            
            crate_nums = [crate for crate in self.crates if crate.startswith("Crate ")]
            crate_rsl = [crate for crate in self.crates if crate.startswith("RSL")]
            crate_nums.sort(key=number)
            crate_rsl.sort()
            self.crates = crate_nums + crate_rsl
            crate_progress = {}
            for crate in self.crates:
                if crate.startswith("Crate"):
                    crate_progress[crate.split()[1]] = self.find_progress(crate.split()[1])
                else:
                    crate_progress[crate] = "PRINT"
            i = 0

            self.columns = {"CRATE": 0, "PRE\nPREP": 1, "  GRADING  ": 2, "  HIGHLIGHTS  ": 3, "   INTRO  ": 4, "RSL /\nPRICE": 5, "PRINT": 6}

            # Styling separater line for corresponding theme
            style = ttk.Style()
            if ctk.get_appearance_mode() == "Dark":
                style.configure("Separator.Horizontal.TSeparator", background="white")
                style.configure("Separator.Vertical.TSeparator", background="white")
            else:
                style.configure("Separator.Horizontal.TSeparator", background="black")
                style.configure("Separator.Vertical.TSeparator", background="black")

            for column, idx in self.columns.items():
                label = ctk.CTkLabel(self.choose_file_frame_scrollable, text=column, font=("Courier New Greek", 18))
                label.grid(row=0, column=idx*2, padx=10, pady=10)

                # Add a horizontal separator after the column title
                h_separator = ttk.Separator(self.choose_file_frame_scrollable, orient='horizontal', style="Separator.Horizontal.TSeparator")
                h_separator.grid(row=1, column=idx*2, columnspan=2, sticky='ew')

                # Add a vertical separator after the column title
                v_separator = ttk.Separator(self.choose_file_frame_scrollable, orient='vertical', style="Separator.Vertical.TSeparator")
                v_separator.grid(row=0, column=idx*2+1, rowspan=i+2, sticky='ns')

            for crate, prog in crate_progress.items():
                txt = str(crate)
                btn = ctk.CTkButton(self.choose_file_frame_scrollable, text=txt, command=lambda file=self.find_file(crate), crate_num=crate: self.create_file_handler(file, crate_num), font=("Courier New Greek", 18), width=50)
                btn.grid(row=i+2, column=0, pady=20, sticky='nsew', padx=10)  # Change row to i+2

                # Add labels of "X" in each column after the "CRATE" column based on the prog

                if prog == "  GRADING  ":
                    # PLACE IN THE COLUMN FOR PREPREP 
                    ctk.CTkLabel(self.choose_file_frame_scrollable, text="\u2713", font=("Courier New Greek", 25)).grid(row=i+2, column=2, pady=10, sticky='nsew', padx=10)

                elif prog == "  HIGHLIGHTS  ":
                    # PLACE IN THE COLUMN FOR PREPREP AND GRADING
                    ctk.CTkLabel(self.choose_file_frame_scrollable, text="\u2713", font=("Courier New Greek", 25)).grid(row=i+2, column=2, pady=10, sticky='nsew', padx=10)

                    ctk.CTkLabel(self.choose_file_frame_scrollable, text="\u2713", font=("Courier New Greek", 25)).grid(row=i+2, column=4, pady=10, sticky='nsew', padx=10)


                elif prog == "   INTRO  ":
                    # PLACE IN THE COLUMN FOR PREPREP, GRADING, AND HIGHLIGHTS
                    ctk.CTkLabel(self.choose_file_frame_scrollable, text="\u2713", font=("Courier New Greek", 25)).grid(row=i+2, column=2, pady=10, sticky='nsew', padx=10)
                    
                    ctk.CTkLabel(self.choose_file_frame_scrollable, text="\u2713", font=("Courier New Greek", 25)).grid(row=i+2, column=4, pady=10, sticky='nsew', padx=10)

                    ctk.CTkLabel(self.choose_file_frame_scrollable, text="\u2713", font=("Courier New Greek", 25)).grid(row=i+2, column=6, pady=10, sticky='nsew', padx=10)



                elif prog == "RSL /\nPRICE":
                    # PLACE IN THE COLUMN FOR PREPREP, GRADING, HIGHLIGHTS, AND INTRO
                    ctk.CTkLabel(self.choose_file_frame_scrollable, text="\u2713", font=("Courier New Greek", 25)).grid(row=i+2, column=2, pady=10, sticky='nsew', padx=10)
                    
                    ctk.CTkLabel(self.choose_file_frame_scrollable, text="\u2713", font=("Courier New Greek", 25)).grid(row=i+2, column=4, pady=10, sticky='nsew', padx=10)

                    ctk.CTkLabel(self.choose_file_frame_scrollable, text="\u2713", font=("Courier New Greek", 25)).grid(row=i+2, column=6, pady=10, sticky='nsew', padx=10)

                    ctk.CTkLabel(self.choose_file_frame_scrollable, text="\u2713", font=("Courier New Greek", 25)).grid(row=i+2, column=8, pady=10, sticky='nsew', padx=10)


                elif prog == "  PRINT  " or txt.startswith("RSL"):
                    print(txt + " : checklist")
                    # PLACE IN THE COLUMN FOR PREPREP, GRADING, HIGHLIGHTS, INTRO, AND RSL /\nPRICE
                    ctk.CTkLabel(self.choose_file_frame_scrollable, text="\u2713", font=("Courier New Greek", 25)).grid(row=i+2, column=2, pady=10, sticky='nsew', padx=10)
                    
                    ctk.CTkLabel(self.choose_file_frame_scrollable, text="\u2713", font=("Courier New Greek", 25)).grid(row=i+2, column=4, pady=10, sticky='nsew', padx=10)

                    ctk.CTkLabel(self.choose_file_frame_scrollable, text="\u2713", font=("Courier New Greek", 25)).grid(row=i+2, column=6, pady=10, sticky='nsew', padx=10)

                    ctk.CTkLabel(self.choose_file_frame_scrollable, text="\u2713", font=("Courier New Greek", 25)).grid(row=i+2, column=8, pady=10, sticky='nsew', padx=10)

                    ctk.CTkLabel(self.choose_file_frame_scrollable, text="\u2713", font=("Courier New Greek", 25)).grid(row=i+2, column=10, pady=10, sticky='nsew', padx=10)


                i += 1
        else:
            self.choose_file_label = ctk.CTkLabel(self.choose_file_frame_scrollable, text="No open crates", font=("Courier New Greek", 18))
            self.choose_file_label.grid(row=0, column=0, columnspan=2, pady=20, sticky='nsew', padx=10)

        

    def create_file_handler(self, file, crate_num):
        self.selected_file = file

        print(str(crate_num) + " selected")
        
        self.title("LP CRATE - " + self.selected_file)
        self.xl_handle = xl.open_excel_file(self.selected_file)
        self.xl_handle_db = self.xl_handle.read_into_dataframe
        self.current_crate_number(crate_num)
                
    def find_file(self, crate_num):

        for file, crates in self.file_crate_dict.items():
            if crate_num.isdigit():
                if "Crate " + str(crate_num) in crates:
                    return file
            else:
                if crate_num in crates:
                    return file
            
    def find_progress(self, crate_num):

        file = self.find_file(crate_num)
        
        excel_file = xl.open_excel_file(file)
                
        needs_hightlights = excel_file.get_row_indexes_needing_highlights(crate_num)
        needs_rsl = excel_file.get_row_indexes_needing_RSL(crate_num)
        needs_grading = excel_file.get_row_indexes_needing_grading(crate_num)
        needs_intro = excel_file.get_row_indexes_needing_intro(crate_num)

        if len(needs_grading) != 0:
            return "  GRADING  "
        elif len(needs_hightlights) != 0:
            return "  HIGHLIGHTS  "
        elif len(needs_intro) != 0:
            return "   INTRO  "
        elif len(needs_rsl) != 0:
            return "RSL /\nPRICE"
        else:
            return "  PRINT  "
        
    def create_new_crate(self):

        self.new_crate = True

        for widget in self.choose_file_frame.winfo_children():
            widget.grid_forget()

        

        files = glob.glob('*')

        xl_handle = min(files, key=os.path.getmtime)

        self.xl_handle = xl.open_excel_file(xl_handle)

        crates_avail = ctk.CTkOptionMenu(self.choose_file_frame, values=[str(num) for num in range(1,21) if "Crate " + str(num) not in self.crates], dropdown_font=("Courier New Greek", 18), font=("Courier New Greek", 18))
        crates_avail.grid(row=0, column=0, pady=20, padx=10)
        
        confirm = ctk.CTkButton(self.choose_file_frame, text="Confirm", command=lambda: self.current_crate_number(crates_avail.get(), True), font=("Courier New Greek", 18))
        confirm.grid(row=0, column=1, pady=20, padx=10)

    def current_crate_number(self, crate_num, new_crate=False):
        for widget in self.choose_file_frame.winfo_children():
            widget.grid_forget()

        if crate_num.startswith("RSL"):
            print("Starts with RSL")
            self.current_crate_num = crate_num
            self.progress_frame = prog.Progress(self, self.current_crate_num)
            self.progress_frame.set_progress("  PRINT  ")
            self.clear()
            self.display_bars()
            self.section6_start(self.current_crate_num)
            
        else:
            print("Does not start with RSL")
            self.current_crate_num = int(crate_num)

            #print("FUNCTION current_crate_number -- crate: " + str(self.current_crate_num))

            self.progress_frame = prog.Progress(self, self.current_crate_num)

            if not new_crate:
                self.is_preprep()
            else:
                self.start_new_vinyl()

    def is_preprep(self):
        for widget in self.choose_file_frame.winfo_children():
            widget.grid_forget()
        label = ctk.CTkLabel(self.choose_file_frame, text="Still Pre-prepping?", font=("Courier New Greek", 18))
        label.grid(row=0, column=0, columnspan=2, pady=20, sticky='nsew', padx=10)

        yes_btn = ctk.CTkButton(self.choose_file_frame, text="Yes", command=self.start_new_vinyl, font=("Courier New Greek", 18))
        yes_btn.grid(row=1, column=0, columnspan=2, pady=20, sticky='nsew', padx=10)

        no_btn = ctk.CTkButton(self.choose_file_frame, text="No", command=self.start, font=("Courier New Greek", 18))
        no_btn.grid(row=2, column=0, columnspan=2, pady=20, sticky='nsew', padx=10)

    def start_new_vinyl(self):
        self.clear()
        self.progress_frame.set_progress("PRE\nPREP")
        self.display_bars()
        self.section1()

    def start(self):
        #print("FUNCTION start -- before clear crate: " + str(self.current_crate_num))
        self.clear()
        #print("FUNCTION start -- before refresh crate: " + str(self.current_crate_num))
        self.xl_handle.refresh_excel_file()
        #print("FUNCTION start -- after refresh crate: " + str(self.current_crate_num))
        self.needs_hightlights = self.xl_handle.get_row_indexes_needing_highlights(self.current_crate_num)
        #print("Indexes that need Highlights: " + str(self.needs_hightlights))
        self.needs_grading = self.xl_handle.get_row_indexes_needing_grading(self.current_crate_num)
        #print("Indexes that need Grading: " + str(self.needs_grading))
        self.needs_rsl = self.xl_handle.get_row_indexes_needing_RSL(self.current_crate_num)
        #print("Indexes that need RSL: " + str(self.needs_rsl))
        self.needs_description = self.xl_handle.get_row_indexes_needing_intro(self.current_crate_num)
        print("Indexes that need INTRO: " + str(self.needs_description))

        if len(self.needs_grading) != 0:
            self.progress_frame.set_progress("  GRADING  ")
            self.display_bars()
            self.section3()
            self.title("GRADING")
        elif len(self.needs_hightlights) != 0:
            self.progress_frame.set_progress("  HIGHLIGHTS  ")
            self.display_bars()
            self.section2()
            self.title("HIGHLIGHTS")
        elif len(self.needs_description) != 0:
            self.progress_frame.set_progress("   INTRO  ")
            self.display_bars()
            self.section4()
            self.title("INTRO")
        elif len(self.needs_rsl) != 0:
            self.progress_frame.set_progress("RSL /\nPRICE")
            self.display_bars()
            self.section5()
            self.title("RSL / PRICE")


    def clear(self):
        for widget in self.winfo_children():
            widget.grid_forget()
        
    def display_bars(self):
        self.progress_frame.grid(row=0, column=0, sticky='nsew', columnspan=7, padx=10, pady=10)
        
    def section1(self):

        crate_index = self.xl_handle.is_crate_there(self.current_crate_num)
        if crate_index == -1:
            self.inputs = dif.DescriptionInputFrame(self)
        else:
            self.inputs = dif.DescriptionInputFrame(self, crate_index)
        self.inputs.reset_fields()
        self.inputs.grid(row=1, column=0, rowspan=20, columnspan=7,sticky="nsew", padx=10, pady=10)

    def section2(self):
        second_sect = II.Frame2(self, self.needs_hightlights, self.xl_handle)
        second_sect.grid(row=1, column=0, rowspan=20, columnspan=7, padx=(10,10), sticky="nsew")



    def section3(self):
        
        self.third_sect = III.Frame2(self, self.needs_grading, self.xl_handle)
        self.third_sect.grid(row=1, column=0, rowspan=20, columnspan=7, padx=(10,10), sticky="nsew")

    def section4(self):
        self.third_sect = III.Intro_Outro(self, self.needs_description, self.xl_handle)
        self.third_sect.grid(row=1, column=0, rowspan=20, columnspan=7, padx=(10,10), sticky="nsew")

    def section5(self):
        dt = datetime.datetime.now()
        rsl_num = "RSL" + dt.strftime("%y%m%d")
        self.rls_frame = rsl.RSL(self, self.needs_rsl, self.xl_handle, rsl_num)
        self.rls_frame.grid(row=1, column=0, rowspan=20, columnspan=7, padx=10, sticky="nsew")
        
    def section6_start(self, ref_num):
        self.clear()
        self.display_bars()
        self.needs_print = self.xl_handle.get_row_indexes_needing_printing(ref_num)
        print("Idxs that need printing: " + str(self.needs_print))
        print("Entering into section6 method")

        self.tree_frame = ctk.CTkScrollableFrame(self, orientation="vertical")
        self.tree_frame.grid(row=1, column=0, columnspan=7, padx=10, pady=15, sticky="nsew")

        self.tree_frame.columnconfigure(0, weight=1)

        # Get the actual column names from self.xl_handle
        columns = self.xl_handle.get_column_names()
        columns.pop(-1)
        columns.pop(-1)

        style = ttk.Style()
        style.configure('MyStyle.Treeview', rowheight=55,
                font=('Helvetica', 12))

        self.treeview = ttk.Treeview(self.tree_frame, columns=columns, show="headings", style='MyStyle.Treeview')
        self.treeview.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=10, pady=10)
        
        self.scrollbar = ttk.Scrollbar(self.tree_frame, orient="horizontal")
        self.scrollbar.grid(row=2, column=0, columnspan=2, sticky="nsew")
        self.treeview.configure(xscrollcommand=self.scrollbar.set)
        self.scrollbar.configure(command=self.treeview.xview)

        # Add column headings
        for col in columns:
            self.treeview.heading(col, text=col)
            if col == "Notes":
                self.treeview.column(col, width=500, stretch= True)
            elif col == "Qty":
                self.treeview.column(col, width=20)
            elif col == "Assignment":
                self.treeview.column(col, width=150)
            else:
                self.treeview.column(col, width=100)

        df = self.xl_handle.read_into_dataframe()

        # Insert the rows of data into the treeview
        for idx in self.needs_print:
            values = [df.loc[idx, col] for col in columns]
            self.treeview.insert("", "end", values=values)

        self.treeview.bind("<Button-1>", self.on_column_click)

        self.treeview.configure(xscrollcommand=self.scrollbar.set)
        self.scrollbar.configure(command=self.treeview.xview)

        self.print_frame = ctk.CTkFrame(self)
        self.print_frame.grid(row=2, column=0, columnspan=7, padx=10, pady=15, sticky="nsew")


        label = ctk.CTkLabel(self.print_frame, text="Ready to print?", font=("Courier New Greek", 18))
        label.grid(row=0, column=0, sticky='nsew', padx=10, pady=10)

        confirm = ctk.CTkButton(self.print_frame, text="Confirm", command=lambda: threading.Thread(target=self.section6_confirm).start(), font=("Courier New Greek", 18))
        confirm.grid(row=0, column=1, sticky='nsew', padx=10, pady=10)
    
    def on_column_click(self, event):
        print(event)
        col = event.widget.identify_column(event.x)
        if col == "#0":
            width = 750
        else:
            width = 100
        event.widget.column(col, width=width+10)

    def section6_confirm(self):
        print("Confirmed")
        self.clear()
        self.progress_frame.set_progress("  Print  ")
        self.display_bars()
        label = ctk.CTkLabel(self, text="Transfering, please wait", font=("Courier New Greek", 18))
        label.grid(row=1, column=0, columnspan=4, sticky='nsew', padx=10, pady=10)

        self.transfer_to_rsl()

    def transfer_to_rsl(self):
        self.source = self.xl_handle.filename
        self.source_df = pd.read_excel(self.source)
#self.destination = "D:\\RSL\\RSL Load.xlsx"
        
        # Load the Excel file
        wb = load_workbook(self.destination)

        # Select the active worksheet
        ws = wb.active

        # Clear all rows in ws after index 1
        ws.delete_rows(2, ws.max_row)

        # Select only the rows that need to be printed
        df_copy = self.source_df.loc[self.needs_print].copy()
        df_copy.drop(columns=["Damages"], inplace=True)
        
        # Save the updated DataFrame to an Excel file, omitting the index
        df_copy.to_excel(self.destination, index=False)

        for idx in self.needs_print:
            row = self.source_df.loc[idx]
            rsl = {}
            rsl["Assignment"] = "P-" + row["Assignment"]
            
            self.xl_handle.update_excel(idx, rsl)

        self.clear()
        self.display_bars()
        self.section6()

    def section6(self):
        
        self.start_app()

    def save_data(self, vinyl_data, row = None, current_section = None):
        self.vinyl_data.update(vinyl_data)
        if current_section == 1:
            self.xl_handle.new_row(self.vinyl_data["Assignment"])
            self.xl_handle.writeOnXL(self.vinyl_data)
        else: 
            self.xl_handle.insert_row(row, self.vinyl_data)

    def update_row(self, row_idx, data):
        self.xl_handle.update_excel(row_idx, data)


if __name__ == "__main__":
    app = Main()
    app.mainloop()